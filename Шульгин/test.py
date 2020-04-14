import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font
from Шульгин.helpers.helpers import get_profile, get_leader_id, get_leader_id_progress

wb2 = load_workbook('итоговая выгрузка.v3 (2).xlsx')
sheets = wb2.sheetnames
ws = wb2[sheets[0]]

questions = [
    "Чему новому, как вам кажется, вы научились ?",
    "Выберите из предложенного списка то, чему вы могли научиться / научились в рамках программы:",
    "Поделитесь вашими комментариями или пожеланиями по поводу того, с чем вы познакомились?",
    "Перечислите основные этапы деятельности, в которую вы были вовлечены и благодаря которой научились чему-то"
]

questions_2 = [
    "Выявлять охраноспособные результаты интеллектуальной деятельности.",
    "Принимать меры по правовой охране результатов интеллектуальной деятельности.",
    "Управлять правами на результаты интеллектуальной деятельности.",
    "Принимать меры по защите интеллектуальных прав в случае их нарушения."
]

res = []

def is_true(val):
    if val.lower() in ["true"]:
        return True
    elif val.lower() in ["false"]:
        return False
    else:
        print("OOOOO", val)

for index in list(range(2, 130)):
    username = ws['R'+str(index)].value
    q_1 = ws['H'+str(index)].value
    if q_1 is not None:
        res.append([username, questions[0], q_1])
    q_3 = ws['M'+str(index)].value
    if q_3 is not None:
        res.append([username, questions[2], q_3])
    q_4 = ws['N'+str(index)].value
    if q_4 is not None:
        res.append([username, questions[3], q_4])
    q_2 = ""
    if is_true(ws['I'+str(index)].value):
        q_2 += questions_2[0]+","
    if is_true(ws['J'+str(index)].value):
        q_2 += questions_2[1]+","
    if is_true(ws['K'+str(index)].value):
        q_2 += questions_2[2]+","
    if is_true(ws['L'+str(index)].value):
        q_2 += questions_2[3]

    if len(q_2) > 0:
        if q_2[-1] == ",":
            q_2 = q_2[:-1]
    # print(q_2)
    if len(q_2) > 0:
        res.append([username, questions[1], q_2])
    #
    # print(ws['H'+str(index)].value)
    # print(ws['M' + str(index)].value)
    # print(ws['N' + str(index)].value)


print("______________________________")
print(res)
print(len(res))

wb_new = Workbook()
ws_new = wb_new.active

for idx, l in enumerate(res):
    ws_new['A'+str(idx+1)] = l[0]
    ws_new['B' + str(idx + 1)] = l[1]
    ws_new['C' + str(idx + 1)] = l[2]

wb_new.save("1 запуск анкеты.xlsx")