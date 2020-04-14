import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
# from helpers.helpers import *
from Шульгин.helpers.helpers import get_leader_id

df = pd.read_excel('reflection_src.xlsx', sheet_name='Лист1')

questions = [
    "Чему новому, как вам кажется, вы научились ?",
    "Выберите из предложенного списка то, чему вы могли научиться / научились в рамках программы:",
    "Поделитесь вашими комментариями или пожеланиями по поводу того, с чем вы познакомились?",
    "Перечислите основные этапы деятельности, в которую вы были вовлечены и благодаря которой научились чему-то"
]

new_answers = {}
answers = df.to_dict('index')

test_list = []
for ind in answers.keys():
    test_list.append(answers[ind]['Username'])
    new_answers[answers[ind]['Username']] = {
        questions[0]: "",
        questions[1]: "",
        questions[2]: "",
        questions[3]: ""
    }

for ind in answers.keys():
    new_answers[answers[ind]['Username']][answers[ind]['Вопрос']] = answers[ind]['Ответ']

final_list = {}
for t in new_answers.keys():
    if len(get_leader_id(t)) > 0:
        final_list[t] = new_answers[t]

print(len(final_list.keys()))
print(final_list)

wb = Workbook()
ws = wb.active

from openpyxl import load_workbook
wb2 = load_workbook('Итог для ЦС.xlsx')
sheets2 = wb2.sheetnames
ws2 = wb2[sheets2[0]]

ws2['F1'] = 'Leader ID'
ws2['F1'].font = Font(bold=True)
ws2['G1'] = 'Дата прохождения рефлексии'
ws2['G1'].font = Font(bold=True)
ws2['H1'] = questions[0]
ws2['H1'].font = Font(bold=True)
ws2['I1'] = questions[1]
ws2['I1'].font = Font(bold=True)
ws2['J1'] = questions[2]
ws2['J1'].font = Font(bold=True)
ws2['K1'] = questions[3]
ws2['K1'].font = Font(bold=True)

# 0ll = 0
for index in list(range(2, 618)):
    if ws2['E'+str(index)].value in final_list.keys():
        # print(index)
        ws2['F' + str(index)] = get_leader_id(ws2['E'+str(index)].value)
        ws2['G' + str(index + 2)] = ""
        ws2['H' + str(index)] = new_answers[ws2['E'+str(index)].value][questions[0]]
        ws2['I' + str(index)] = new_answers[ws2['E'+str(index)].value][questions[1]]
        ws2['J' + str(index)] = new_answers[ws2['E'+str(index)].value][questions[2]]
        ws2['K' + str(index)] = new_answers[ws2['E'+str(index)].value][questions[3]]
        # print(ws2['E'+str(index)].value)
        # ll+=1
# print(ll)


#
# for index, l in enumerate(final_list.keys()):
#     ws['A'+str(index+2)] = get_leader_id(l)
#     ws['B' + str(index + 2)] = ""
#     ws['C'+str(index+2)] = new_answers[l][questions[0]]
#     ws['D'+str(index+2)] = new_answers[l][questions[1]]
#     ws['E'+str(index+2)] = new_answers[l][questions[2]]
#     ws['F'+str(index+2)] = new_answers[l][questions[3]]
    # ws['G'+str(index+2)] = l


wb2.save("reflection_2035.xlsx")

# -------------------------------------------------------

